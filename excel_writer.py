from openpyxl import load_workbook
from datetime import datetime
import os

EXCEL_PATH = "bp_records.xlsx"   # 你的 Excel 檔名
START_ROW = 52                   # 從第 52 行開始
STUDENT_ID = "C112252224"

def write_bp_to_excel(sys, dia):
    # 如果 Excel 不存在 → 建立一個
    if not os.path.exists(EXCEL_PATH):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # 找到下一個空白列（至少從 52 行開始）
    row = START_ROW
    while ws[f"A{row}"].value is not None:
        row += 1

    # 時間格式
    now = datetime.now().strftime("%Y/%m/%d %I:%M %p")

    # 寫入資料
    ws[f"A{row}"] = STUDENT_ID
    ws[f"B{row}"] = now
    ws[f"C{row}"] = sys
    ws[f"D{row}"] = dia

    wb.save(EXCEL_PATH)
